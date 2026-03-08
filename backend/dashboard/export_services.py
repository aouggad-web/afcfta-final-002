"""
AfCFTA Platform - Export Services
===================================
Generates PDF and Excel reports with executive summaries.
Falls back gracefully if optional libraries (reportlab, openpyxl) are absent.
"""

from __future__ import annotations

import io
import json
import logging
import os
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class ExportService:
    """
    Exports analytics reports to PDF and Excel formats.
    """

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    def export_investment_analysis_excel(
        self,
        analysis_data: Dict[str, Any],
        filename: Optional[str] = None,
    ) -> bytes:
        """
        Export investment analysis results to Excel (.xlsx).
        Returns raw bytes of the workbook.
        """
        if not OPENPYXL_AVAILABLE:
            return self._csv_fallback(analysis_data)

        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill("solid", fgColor="1A365D")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # Title
        ws_summary["A1"] = "AfCFTA Investment Intelligence Report"
        ws_summary["A1"].font = Font(bold=True, size=16, color="1A365D")
        ws_summary.merge_cells("A1:F1")
        ws_summary["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        ws_summary["A2"].font = Font(italic=True, color="666666")
        ws_summary.merge_cells("A2:F2")

        # Opportunity table headers
        row = 4
        headers = ["Rank", "Country", "Sector", "Score (%)", "Grade", "Recommendation"]
        for col, h in enumerate(headers, start=1):
            cell = ws_summary.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Data rows
        opportunities = analysis_data.get("opportunities", [])
        for i, opp in enumerate(opportunities[:50], start=1):
            data_row = [
                i,
                opp.get("country", ""),
                opp.get("sector", ""),
                round(float(opp.get("investment_score", 0)) * 100, 1),
                opp.get("grade", ""),
                opp.get("recommendation", "").replace("_", " ").title(),
            ]
            for col, value in enumerate(data_row, start=1):
                cell = ws_summary.cell(row=row + i, column=col, value=value)
                cell.border = thin_border
                if col == 4:  # Score column
                    score = float(opp.get("investment_score", 0))
                    if score >= 0.75:
                        cell.fill = PatternFill("solid", fgColor="D1FAE5")
                    elif score >= 0.60:
                        cell.fill = PatternFill("solid", fgColor="DBEAFE")
                    elif score < 0.45:
                        cell.fill = PatternFill("solid", fgColor="FEE2E2")

        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws_summary.column_dimensions[get_column_letter(col)].width = 18

        # Regional summary sheet
        ws_regional = wb.create_sheet("Regional Analysis")
        ws_regional["A1"] = "Regional Bloc Benchmarks"
        ws_regional["A1"].font = Font(bold=True, size=14, color="1A365D")

        try:
            from intelligence.analytics.regional_analytics import REGIONAL_BENCHMARKS
            reg_headers = ["Bloc", "Avg Tariff (%)", "Investment Score", "Infrastructure", "Trade Integration (%)"]
            for col, h in enumerate(reg_headers, start=1):
                cell = ws_regional.cell(row=3, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2D3748")
                cell.alignment = Alignment(horizontal="center")

            for row_idx, (bloc, metrics) in enumerate(REGIONAL_BENCHMARKS.items(), start=4):
                ws_regional.cell(row=row_idx, column=1, value=bloc)
                ws_regional.cell(row=row_idx, column=2, value=metrics.get("tariff_avg", ""))
                ws_regional.cell(row=row_idx, column=3, value=round(metrics.get("investment_score", 0) * 100, 1))
                ws_regional.cell(row=row_idx, column=4, value=round(metrics.get("infrastructure", 0) * 100, 1))
                ws_regional.cell(row=row_idx, column=5, value=metrics.get("trade_integration", ""))

            for col in range(1, len(reg_headers) + 1):
                ws_regional.column_dimensions[get_column_letter(col)].width = 22
        except Exception as exc:
            logger.warning(f"Regional data unavailable for Excel export: {exc}")

        # Serialise to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ------------------------------------------------------------------
    # PDF export
    # ------------------------------------------------------------------

    def export_investment_analysis_pdf(
        self,
        analysis_data: Dict[str, Any],
    ) -> bytes:
        """
        Export investment analysis results to PDF with executive summary.
        Returns raw bytes of the PDF document.
        """
        if not REPORTLAB_AVAILABLE:
            return self._json_fallback(analysis_data)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = styles["Title"]
        story.append(Paragraph("AfCFTA Investment Intelligence Report", title_style))
        story.append(Paragraph(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            styles["Normal"]
        ))
        story.append(Spacer(1, 20))

        # Executive Summary
        story.append(Paragraph("Executive Summary", styles["Heading1"]))
        opp_count = analysis_data.get("total_count", len(analysis_data.get("opportunities", [])))
        story.append(Paragraph(
            f"This report analysed {opp_count} investment opportunities across African markets "
            "using the AfCFTA AI Investment Intelligence Engine. Opportunities are ranked by "
            "composite investment score incorporating market access, investment climate, "
            "infrastructure quality, incentive packages, market potential, and cost competitiveness.",
            styles["Normal"]
        ))
        story.append(Spacer(1, 12))

        # Top opportunities table
        story.append(Paragraph("Top Investment Opportunities", styles["Heading1"]))
        opportunities = analysis_data.get("opportunities", [])[:15]

        table_data = [["#", "Country", "Sector", "Score", "Grade"]]
        for i, opp in enumerate(opportunities, start=1):
            table_data.append([
                str(i),
                opp.get("country", ""),
                opp.get("sector", "general").title(),
                f"{float(opp.get('investment_score', 0)) * 100:.1f}%",
                opp.get("grade", ""),
            ])

        table = Table(table_data, colWidths=[30, 80, 100, 60, 40])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A365D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFC")]),
        ]))
        story.append(table)

        doc.build(story)
        return buffer.getvalue()

    # ------------------------------------------------------------------
    # Fallbacks
    # ------------------------------------------------------------------

    @staticmethod
    def _csv_fallback(data: Dict[str, Any]) -> bytes:
        """Return a CSV bytes stream when openpyxl is unavailable."""
        lines = ["Country,Sector,Score,Grade,Recommendation"]
        for opp in data.get("opportunities", []):
            lines.append(
                f"{opp.get('country','')},{opp.get('sector','')},"
                f"{opp.get('investment_score',0):.3f},{opp.get('grade','')},"
                f"{opp.get('recommendation','')}"
            )
        return "\n".join(lines).encode("utf-8")

    @staticmethod
    def _json_fallback(data: Dict[str, Any]) -> bytes:
        """Return a JSON bytes stream when reportlab is unavailable."""
        return json.dumps(data, default=str, indent=2).encode("utf-8")


# Singleton
_service: Optional[ExportService] = None


def get_export_service() -> ExportService:
    global _service
    if _service is None:
        _service = ExportService()
    return _service
